from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Assets" / "ORF Fixes Pronunciations.xlsx"
OUTPUT = ROOT / "Assets" / "orf-fix-pronunciations.json"


def cell_text(value: object) -> str:
    return str(value or "").strip()


def main() -> None:
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("Pronunciation workbook is empty.")

    headers = [cell_text(cell).lower() for cell in rows[0]]
    required = ["ident", "spoken", "aliases"]
    missing = [header for header in required if header not in headers]
    if missing:
        raise SystemExit(f"Missing required columns: {', '.join(missing)}")

    ident_index = headers.index("ident")
    spoken_index = headers.index("spoken")
    aliases_index = headers.index("aliases")
    pronunciations: dict[str, dict[str, object]] = {}

    for row in rows[1:]:
        ident = cell_text(row[ident_index]).upper()
        if not ident:
            continue
        spoken = cell_text(row[spoken_index]) or ident
        alias_text = cell_text(row[aliases_index])
        aliases = [alias.strip() for alias in alias_text.split(",") if alias.strip()]
        pronunciations[ident] = {
            "spoken": spoken,
            "aliases": aliases,
        }

    OUTPUT.write_text(json.dumps(pronunciations, indent=2, sort_keys=True) + "\n")
    print(f"Wrote {len(pronunciations)} fix pronunciations to {OUTPUT}")


if __name__ == "__main__":
    main()
