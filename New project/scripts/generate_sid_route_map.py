from __future__ import annotations

import json
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # The dev server may run from a Python without openpyxl.
    load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Assets" / "ORF_Destination_SID_Transition_Route_Map.xlsx"
OUTPUT = ROOT / "Assets" / "orf-sid-route-map.json"


def cell_text(value: object) -> str:
    return str(value or "").strip()


def rows_from_sheet(workbook, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [cell_text(cell).lower() for cell in rows[0]]
    output: list[dict[str, str]] = []
    for row in rows[1:]:
        record = {
            headers[index]: cell_text(value)
            for index, value in enumerate(row[: len(headers)])
            if headers[index]
        }
        if any(record.values()):
            output.append(record)
    return output


def column_index(cell_ref: str) -> int:
    letters = re.match(r"([A-Z]+)", cell_ref or "")
    if not letters:
        return 0
    index = 0
    for char in letters.group(1):
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def text_from_rich_text(node: ET.Element) -> str:
    return "".join(text.text or "" for text in node.iter() if text.tag.endswith("}t"))


def xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    return [text_from_rich_text(item) for item in root]


def xlsx_sheet_paths(archive: zipfile.ZipFile) -> dict[str, str]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("pkg:Relationship", ns)
    }
    paths = {}
    for sheet in workbook.findall("main:sheets/main:sheet", ns):
        rel_id = sheet.attrib.get(f"{{{ns['rel']}}}id", "")
        target = targets.get(rel_id, "")
        if target:
            clean_target = target.lstrip("/")
            paths[sheet.attrib.get("name", "")] = clean_target if clean_target.startswith("xl/") else f"xl/{clean_target}"
    return paths


def rows_from_xlsx(path: Path, sheet_name: str) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = xlsx_shared_strings(archive)
        sheet_path = xlsx_sheet_paths(archive).get(sheet_name)
        if not sheet_path:
            return []
        root = ET.fromstring(archive.read(sheet_path))
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    raw_rows: list[list[str]] = []
    for row in root.findall(".//main:sheetData/main:row", ns):
        cells: dict[int, str] = {}
        for cell in row.findall("main:c", ns):
            index = column_index(cell.attrib.get("r", ""))
            cell_type = cell.attrib.get("t", "")
            value_node = cell.find("main:v", ns)
            if cell_type == "inlineStr":
                value = text_from_rich_text(cell)
            elif value_node is None:
                value = ""
            elif cell_type == "s":
                value = shared_strings[int(value_node.text or 0)]
            else:
                value = value_node.text or ""
            cells[index] = cell_text(value)
        if cells:
            raw_rows.append([cells.get(index, "") for index in range(max(cells) + 1)])
    if not raw_rows:
        return []
    headers = [cell_text(cell).lower() for cell in raw_rows[0]]
    output: list[dict[str, str]] = []
    for row in raw_rows[1:]:
        record = {
            headers[index]: cell_text(value)
            for index, value in enumerate(row[: len(headers)])
            if headers[index]
        }
        if any(record.values()):
            output.append(record)
    return output


def main() -> None:
    if load_workbook:
        workbook = load_workbook(SOURCE, read_only=True, data_only=True)
        destination_rows = rows_from_sheet(workbook, "Destination_SID_Transitions")
        transition_rows = rows_from_sheet(workbook, "SID_Transition_Key")
    else:
        destination_rows = rows_from_xlsx(SOURCE, "Destination_SID_Transitions")
        transition_rows = rows_from_xlsx(SOURCE, "SID_Transition_Key")

    transitions = {}
    for row in transition_rows:
        key = row.get("sid_transition_format", "").upper()
        if key:
            transitions[key] = row.get("meaning", "")

    routes = []
    for row in destination_rows:
        sid_transition = row.get("sid_transition", "").upper()
        sid, _, transition = sid_transition.partition(".")
        destinations = [
            destination.strip().upper()
            for destination in row.get("destination_airport", "").replace("/", ",").split(",")
            if destination.strip()
        ]
        routes.append({
            "destinations": destinations,
            "sid_transition": sid_transition,
            "sid": sid,
            "transition": transition,
            "route": row.get("route", ""),
            "alt": row.get("alt", ""),
            "aircraft_type": row.get("type", ""),
            "notes": row.get("notes", ""),
            "meaning": transitions.get(sid_transition, ""),
        })

    payload = {
        "source": SOURCE.name,
        "routes": routes,
        "transition_key": transitions,
    }
    OUTPUT.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
    print(f"Wrote {len(routes)} SID destination routes to {OUTPUT}")


if __name__ == "__main__":
    main()
