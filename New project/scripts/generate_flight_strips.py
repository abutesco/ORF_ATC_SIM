#!/usr/bin/env python3
"""Convert ORF departure and overflight strip workbook into a browser JSON asset."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Assets" / "ORF_Flight_Strips_Departures_Overflights.xlsx"
OUTPUT = ROOT / "Assets" / "orf-flight-strips.json"


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def rows_from_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    headers = [clean(value).lower() for value in header_row]
    rows = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        row = {}
        for header, value in zip(headers, raw_row):
            if header:
                row[header] = clean(value)
        if any(row.values()):
            rows.append(row)
    return rows


def normalize_callsign(value):
    callsign = clean(value).upper()
    if callsign.startswith("NAVY"):
        return "VV" + callsign[4:]
    return callsign


def main():
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    departures = rows_from_sheet(workbook, "Departures_Strips")
    overflights = rows_from_sheet(workbook, "Overflights_Strips")
    for row in overflights:
        row["sample_callsign"] = normalize_callsign(row.get("sample_callsign"))
        row["strip_format"] = clean(row.get("strip_format")).replace("NAVY", "VV")

    data = {
        "source": SOURCE.name,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "departures": departures,
        "overflights": overflights,
        "altitude_rules": rows_from_sheet(workbook, "Altitude_Rules"),
        "aircraft_class_key": rows_from_sheet(workbook, "Aircraft_Class_Key"),
    }
    OUTPUT.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(departures)} departure strips and {len(overflights)} overflight strips to {OUTPUT}")


if __name__ == "__main__":
    main()
