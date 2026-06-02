#!/usr/bin/env python3
"""Convert the editable Norfolk traffic workbook into a browser JSON asset."""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Assets" / "Norfolk_Airspace_Realistic_Traffic_Generator.xlsx"
OUTPUT = ROOT / "Assets" / "norfolk-realistic-traffic.json"


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def normalize_callsign(value):
    callsign = clean(value).upper()
    if callsign.startswith("NAVY"):
        return "VV" + callsign[4:]
    return callsign


def rows_from_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    headers = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        row = {}
        for header, value in zip(headers, raw_row):
            if not header:
                continue
            row[header] = clean(value)
        if any(row.values()):
            rows.append(row)
    return rows


def main():
    workbook = load_workbook(SOURCE, data_only=True)
    traffic = []
    by_airport = defaultdict(list)

    for row in rows_from_sheet(workbook, "Traffic_Generator"):
        airport_ident = clean(row.get("airport_ident")).upper()
        callsign = normalize_callsign(row.get("sample_callsign"))
        aircraft_type = clean(row.get("aircraft_type")).upper()
        if not airport_ident or not callsign or not aircraft_type:
            continue
        entry = {
            "airport_ident": airport_ident,
            "airport_name": clean(row.get("airport_name")),
            "airport_category": clean(row.get("airport_category")),
            "traffic_mix": clean(row.get("traffic_mix")),
            "frequency": clean(row.get("frequency")),
            "callsign": callsign,
            "aircraft_type": aircraft_type,
            "traffic_role": clean(row.get("traffic_role")),
            "use_in_sim": clean(row.get("use_in_sim")),
            "notes": clean(row.get("notes")),
        }
        traffic.append(entry)
        by_airport[airport_ident].append(entry)

    data = {
        "source": SOURCE.name,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "traffic": traffic,
        "by_airport": {ident: rows for ident, rows in sorted(by_airport.items())},
        "airport_summary": rows_from_sheet(workbook, "Airport_Summary"),
        "aircraft_type_key": rows_from_sheet(workbook, "Aircraft_Type_Key"),
    }
    OUTPUT.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(traffic)} traffic rows to {OUTPUT}")


if __name__ == "__main__":
    main()
